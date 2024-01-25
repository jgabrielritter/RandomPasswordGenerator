import random
import string
from datetime import datetime
import openpyxl
from pathlib import Path

# Excel file to store generated passwords
PASSWORDS_FILE = Path.home() / "Desktop" / "PasswordsGenerated.xlsx"

def generate_password():
    # Function to generate a random password
    length = 12
    chars = string.ascii_letters + string.digits + string.punctuation
    return ''.join(random.choice(chars) for _ in range(length))

def load_passwords():
    # Function to load passwords from the Excel file
    try:
        workbook = openpyxl.load_workbook(PASSWORDS_FILE)
        sheet = workbook.active
        passwords = []
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Start from the second row to skip headers
            passwords.append({"#": row[0], "password": row[1], "generated_at": row[2]})
        return passwords
    except FileNotFoundError:
        # If the file doesn't exist, return an empty list
        return []

def save_passwords(passwords):
    # Function to save passwords to the Excel file
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["#", "Password", "Generated At"])  # Write headers only once
    for index, entry in enumerate(passwords, start=1):
        sheet.append([index, entry["password"], entry["generated_at"]])
    workbook.save(PASSWORDS_FILE)

def store_password():
    # Function to generate a password, store it along with the timestamp, and save to Excel file
    passwords = load_passwords()  # Load existing passwords
    password = generate_password()  # Generate a new password
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Append the new password and timestamp to the list
    passwords.append({"#": None, "password": password, "generated_at": current_time})
    
    save_passwords(passwords)  # Save the updated list to the Excel file
    print(f"Password '{password}' generated and stored at {current_time}")

def lookup_passwords():
    # Function to load and print all stored passwords from the Excel file
    passwords = load_passwords()  # Load existing passwords
    for entry in passwords:
        # Print each password along with its number and timestamp
        print(f"#{entry['#']} Password: {entry['password']} (Generated at: {entry['generated_at']})")

# Example usage
store_password()  # Generate and store a new password
lookup_passwords()  # Print all stored passwords
